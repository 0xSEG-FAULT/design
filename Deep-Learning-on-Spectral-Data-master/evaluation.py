import os
import joblib
import math
import torch
import json
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

from constants import *
from training import models_to_train
from utils import save_to_excel, read_feature_list_from_csv
from sklearn.metrics import mean_squared_error, r2_score
from scipy.stats import pearsonr
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from matplotlib.colors import LinearSegmentedColormap, TwoSlopeNorm, Normalize, to_hex
from tensorflow.keras.models import load_model
from transformer_training import SpectraTransformer, DEVICE, apply_spectral_preprocessing
from torch.utils.data import DataLoader, TensorDataset


def get_scaled_predictions_from_transformer(scaled_data, models_dir, model_name):
    with open(f"{models_dir}/best_tuned_parameters.json", "r") as f:
        params = json.load(f)
    model = SpectraTransformer(
        input_size=params["input_size"],
        reduced_size=params["reduced_size"],
        input_channels=params["input_channels"],
        d_model=params["d_model"],
        patch_size=params["patch_size"],
        stride=params["stride"],
        num_layers=params["num_layers"],
        nhead=params["nhead"],
        dim_feedforward=params["dim_feedforward"],
        num_targets=params["num_targets"],
    ).to(DEVICE)
    model.load_state_dict(
        torch.load(f"{models_dir}/{model_name}.pt", weights_only=True)
    )
    scaled_data = torch.from_numpy(scaled_data).float().to(DEVICE)
    scaled_data = TensorDataset(scaled_data)

    dataloader = DataLoader(scaled_data, batch_size=params["batch_size"], shuffle=False)
    predictions = []
    model.eval()
    with torch.no_grad():
        for X_batch in dataloader:
            X_batch = X_batch[0]
            pred_batch = model(X_batch)
            # Convert to numpy for calculations of metrics
            predictions.append(pred_batch.cpu().numpy())

    return np.concatenate(predictions, axis=0)


def load_and_predict(data, models_dir, models, target_names, model_type=CLASSIC_MODEL):
    """
    :return: Dictionary of DataFrames with predictions for each target variable
    """
    predictions = {target: {} for target in target_names}
    for model_name in models.keys():
        if model_type == KERAS_MODEL or model_type == TRANSFORMER_MODEL:
            feature_scaler = joblib.load(f"{models_dir}/feature_scaler.pkl")
            print("Preprocessing...")
            data, pca = apply_spectral_preprocessing(data, apply_pca=False)
            print("Scaling and predicting...")
            scaled_data = feature_scaler.transform(data)
            if model_type == TRANSFORMER_MODEL:
                model_predictions = get_scaled_predictions_from_transformer(
                    scaled_data, models_dir, model_name
                )
            else:
                model = load_model(f"{models_dir}/{model_name}.keras")
                model_predictions = model.predict(scaled_data)
            if os.path.exists(f"{models_dir}/target_scaler.pkl"):
                target_scaler = joblib.load(f"{models_dir}/target_scaler.pkl")
                model_predictions = target_scaler.inverse_transform(model_predictions)
        elif model_type == CLASSIC_MODEL:
            if not os.path.exists(f"{models_dir}/{model_name}.pkl"):
                continue
            model = joblib.load(f"{models_dir}/{model_name}.pkl")
            model_predictions = model.predict(data)
        else:
            raise ValueError("Incorrect model type. See constants.py for valid values")
        if model_predictions.ndim == 1:
            model_predictions = model_predictions[:, np.newaxis]
        for i, target in enumerate(target_names):
            predictions[target][model_name] = model_predictions[:, i]

    for target in target_names:
        # Maintain original index
        predictions[target] = pd.DataFrame(predictions[target], index=data.index)

    print("Predictions ready")
    return predictions


def plot_scatter(predictions, target, plot_dir, n_cols=4):
    n_rows = math.ceil(len(predictions.columns) / n_cols)
    plt.figure(figsize=(n_cols * 4, n_rows * 4))
    for i, model_name in enumerate(predictions.columns):
        plt.subplot(n_rows, n_cols, i + 1)
        plt.scatter(target, predictions[model_name], alpha=0.5, label=model_name)
        plt.plot(
            [target.min(), target.max()], [target.min(), target.max()], "k--", lw=2
        )
        plt.xlim(target.min(), target.max())
        plt.ylim(target.min(), target.max())
        plt.xlabel("True Values", fontsize=14)
        plt.ylabel("Predictions", fontsize=14)
        plt.suptitle(f"Scatter Plot", fontsize=16, weight="bold")
        plt.xticks(fontsize=14)
        plt.yticks(fontsize=14)
        plt.legend()
        plt.grid(alpha=0.3)
        plt.tight_layout()
    plt.savefig(f"{plot_dir}/scatter_plot_{target.name}.png")
    plt.close()
    print(f"Scatter plot saved to {plot_dir}/scatter_plot_{target.name}.png")


def color_sheet_cells(df, worksheet):
    cmap = LinearSegmentedColormap.from_list(
        "excel_gry", ["#63BE7B", "#FFEB84", "#F8696B"]
    )
    cmap_r = LinearSegmentedColormap.from_list(
        "rev_excel_gry", ["#F8696B", "#FFEB84", "#63BE7B"]
    )

    for col_idx, df_col in enumerate(df.columns, start=2):
        if df_col in [
            "Mean Abs Bias",
            "Median Abs Bias",
            "Mean % Bias",
            "Median % Bias",
        ]:
            values = df[df_col].abs().values.flatten()
        else:
            values = df[df_col].values.flatten()
        # NaN values may cause percentile issues
        values = values[~np.isnan(values.astype(float))]
        vmin, vcenter, vmax = np.percentile(values, [0, 50, 100])
        if vmin == vmax:
            norm = Normalize(vmin=vmin, vmax=vmax)
        else:
            if vcenter == vmin or vcenter == vmax:
                # if median is too close to extremes, then use mean
                vcenter = np.mean([vmin, vmax])
            norm = TwoSlopeNorm(vmin=vmin, vcenter=vcenter, vmax=vmax)

        # Write data and color cells
        for row_idx, value in enumerate(df[df_col].values, start=2):
            if df_col in [
                "Mean Abs Bias",
                "Median Abs Bias",
                "Mean % Bias",
                "Median % Bias",
            ]:
                norm_val = norm(abs(value))
            else:
                norm_val = norm(value)
            if df_col in ["R2", "Train R2", "Pearson R", "Train Pearson R"]:
                color_map = cmap_r(norm_val)
            else:
                color_map = cmap(norm_val)
            color = to_hex(color_map, keep_alpha=False).replace("#", "").upper()

            # Set value and background fill
            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )


def evaluate_models(
    training_data,
    test_data,
    metrics_output_file,
    predictions_output_file,
    plots_dir,
    models_dir=MODELS_ROOT,
    models_to_evaluate=models_to_train,
    selected_features_file=None,
    model_type=CLASSIC_MODEL,
):
    env_features_train, spectral_features_train, target_train = training_data
    env_features_test, spectral_features_test, target_test = test_data
    selected_features = None
    if selected_features_file:
        selected_features = read_feature_list_from_csv(selected_features_file)

    features = pd.concat([env_features_train, spectral_features_train], axis=1)
    if selected_features:
        features = features[selected_features]
    # Scalar doesn't support mixed types for column names
    if hasattr(features, "columns"):
        features.columns = [f"feature_{i}" for i in range(features.shape[1])]
    if isinstance(target_train, pd.Series):
        target_train = target_train.to_frame()
        target_test = target_test.to_frame()
    predictions_train = load_and_predict(
        features,
        models_dir,
        models_to_evaluate,
        target_names=target_train.columns,
        model_type=model_type,
    )
    features = pd.concat([env_features_test, spectral_features_test], axis=1)
    if selected_features:
        features = features[selected_features]
    # Scalar doesn't support mixed types for column names
    if hasattr(features, "columns"):
        features.columns = [f"feature_{i}" for i in range(features.shape[1])]
    predictions_test = load_and_predict(
        features,
        models_dir,
        models_to_evaluate,
        target_names=target_test.columns,
        model_type=model_type,
    )

    if not os.path.exists(CLEANED_TEST_DATA_FILE):
        save_to_excel(
            env_features_test,
            spectral_features_test,
            target_test,
            file_path=CLEANED_TEST_DATA_FILE,
        )

    for target in target_test.columns:
        save_to_excel(
            predictions_test[target],
            file_path=predictions_output_file,
            sheet_name=target,
        )

    for target in target_test.columns:
        metrics = pd.DataFrame(
            columns=[
                "Train MSE",
                "Train R2",
                "Train Pearson R",
                "MSE",
                "R2",
                "Pearson R",
                "Mean Abs Bias",
                "Median Abs Bias",
                "Mean % Bias",
                "Median % Bias",
                "Abs Bias Std",
                "% Bias Std",
            ],
            index=models_to_evaluate.keys(),
        )

        for model_name in models_to_evaluate.keys():
            # Correlation metrics
            if len(target_train) >= 2:
                metrics.loc[model_name, "Train Pearson R"] = round(
                    pearsonr(
                        target_train[target], predictions_train[target][model_name]
                    )[0],
                    RESULTS_DECIMAL,
                )
                metrics.loc[model_name, "Train R2"] = round(
                    r2_score(
                        target_train[target], predictions_train[target][model_name]
                    ),
                    RESULTS_DECIMAL,
                )
            else:
                print(
                    f"Too few samples ({len(target_train)}) to calculate Pearson and R2"
                )
                metrics.loc[model_name, "Train Pearson R"] = [np.nan]
                metrics.loc[model_name, "Train R2"] = [np.nan]
            metrics.loc[model_name, "Train MSE"] = round(
                mean_squared_error(
                    target_train[target], predictions_train[target][model_name]
                ),
                RESULTS_DECIMAL,
            )

            if len(target_test) >= 2:
                metrics.loc[model_name, "Pearson R"] = round(
                    pearsonr(target_test[target], predictions_test[target][model_name])[
                        0
                    ],
                    RESULTS_DECIMAL,
                )
                metrics.loc[model_name, "R2"] = round(
                    r2_score(target_test[target], predictions_test[target][model_name]),
                    RESULTS_DECIMAL,
                )
            else:
                print(
                    f"Too few samples ({len(target_test)}) to calculate Pearson and R2"
                )
                metrics.loc[model_name, "Pearson R"] = [np.nan]
                metrics.loc[model_name, "R2"] = [np.nan]
            metrics.loc[model_name, "MSE"] = round(
                mean_squared_error(
                    target_test[target], predictions_test[target][model_name]
                ),
                RESULTS_DECIMAL,
            )

            # Bias metrics
            mean_values = (
                target_test[target] + predictions_test[target][model_name]
            ) / 2
            absolute_diff = predictions_test[target][model_name] - target_test[target]
            percentage_diff = 100 * absolute_diff / mean_values
            # Ignore points where mean becomes zero
            percentage_diff = pd.Series(filter(math.isfinite, percentage_diff))
            metrics.loc[model_name, "Mean Abs Bias"] = round(
                absolute_diff.abs().mean(), RESULTS_DECIMAL
            )
            metrics.loc[model_name, "Median Abs Bias"] = round(
                absolute_diff.abs().median(), RESULTS_DECIMAL
            )
            metrics.loc[model_name, "Mean % Bias"] = round(
                percentage_diff.mean(), RESULTS_DECIMAL
            )
            metrics.loc[model_name, "Median % Bias"] = round(
                percentage_diff.median(), RESULTS_DECIMAL
            )
            metrics.loc[model_name, "Abs Bias Std"] = round(
                absolute_diff.std(), RESULTS_DECIMAL
            )
            metrics.loc[model_name, "% Bias Std"] = round(
                percentage_diff.std(), RESULTS_DECIMAL
            )

        if os.path.exists(metrics_output_file):
            with pd.ExcelWriter(
                metrics_output_file, mode="a", if_sheet_exists="new"
            ) as writer:
                metrics.to_excel(
                    writer, sheet_name=f"{target} Test Metrics", index=True
                )
                metrics[
                    [
                        "Train MSE",
                        "MSE",
                        "Train R2",
                        "R2",
                        "Train Pearson R",
                        "Pearson R",
                    ]
                ].to_excel(writer, sheet_name=f"{target} Train vs Test", index=True)
        else:
            with pd.ExcelWriter(metrics_output_file) as writer:
                metrics.to_excel(
                    writer, sheet_name=f"{target} Test Metrics", index=True
                )
                metrics[
                    [
                        "Train MSE",
                        "MSE",
                        "Train R2",
                        "R2",
                        "Train Pearson R",
                        "Pearson R",
                    ]
                ].to_excel(writer, sheet_name=f"{target} Train vs Test", index=True)
        print(f"Metrics saved to {metrics_output_file}")

        wb = load_workbook(metrics_output_file)
        color_sheet_cells(metrics, wb[f"{target} Test Metrics"])
        color_sheet_cells(
            metrics[
                ["Train MSE", "MSE", "Train R2", "R2", "Train Pearson R", "Pearson R"]
            ],
            wb[f"{target} Train vs Test"],
        )
        wb.save(metrics_output_file)

        # Scatter plot
        if PLOTTING_ENABLED:
            os.makedirs(plots_dir, exist_ok=True)
            plot_scatter(
                predictions_test[target],
                target_test[target],
                plots_dir,
                n_cols=min(len(predictions_test[target].columns), 4),
            )
